import pandas as pd
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import os

# Configuration variables
excel_path = r"O:\NIA\PORTFOLIOS TEAM\GE Analytics\6. Power BI\4. PBI Cleint Projects\UCL RCA\UCL Data.xlsx"
sheet_name = 'UCL Summary'
image_column = 'R' # Column containing the photos
id_column = 'B'    # Column containing names or IDs to name the photo files
github_user = 'H1470'
github_repo = 'KWphotos'
branch = 'main'
folder = 'Images'  # Local folder to save images

os.makedirs(folder, exist_ok=True)

# Load workbook and image loader
wb = openpyxl.load_workbook(excel_path)
sheet = wb[sheet_name]
image_loader = SheetImageLoader(sheet) # Maps images to their specific cells

output_data = []

# Iterate through the rows to extract and map
print(f"Starting to iterate from row 2 to {sheet.max_row}...")
images_found = 0
rows_with_id = 0

for row in range(2, sheet.max_row + 1):
    id_val = str(sheet[f"{id_column}{row}"].value)
    img_cell = f"{image_column}{row}"
    
    if id_val != "None":
        rows_with_id += 1
    
    if id_val != "None" and image_loader.image_in(img_cell):
        images_found += 1
        # Extract and save image locally
        image = image_loader.get(img_cell)
        # Sanitize filename: remove invalid Windows characters and whitespace
        invalid_chars = '<>:"/\\|?*\n\r\t'
        safe_id = id_val.strip()  # Remove leading/trailing whitespace
        for char in invalid_chars:
            safe_id = safe_id.replace(char, '_')
        filename = f"{safe_id}.png"
        image.save(f"{folder}/{filename}")
        
        # Construct the raw GitHub URL predictably
        raw_url = f"https://raw.githubusercontent.com/{github_user}/{github_repo}/{branch}/{folder}/{filename}"
        output_data.append({"Property ID": id_val, "Raw GitHub Link": raw_url})

print(f"\nDebug info:")
print(f"  Rows with ID in column B: {rows_with_id}")
print(f"  Images found and processed: {images_found}")

# Export the ready-to-paste URLs
pd.DataFrame(output_data).to_excel("github_image_links.xlsx", index=False)
print("Extraction and link generation complete!")